
# coding: utf-8

import subprocess as sb
import openpyxl
import psutil
import shutil
import os
import time
import csv
import logging
import xml.dom.minidom
import threading
import pandas as pd
from datetime import datetime,timedelta
from dateutil.parser import parse
import sys

logging.basicConfig(filename = "D:\\Script_Logs\\Test.logs", level = logging.DEBUG, format = '\n%(asctime)s : [%(levelname)s] : %(message)s')
#0 to set root path,1 to write and save results,2 to read and save results,3 to Mixed mode and save results,4 to remove storage
#5 to add delay #6 to replace files
root_path=""
mem=0.0
finish_tracking=1
wb=0
testdict={}
report=""

## to get system parameters using python
##def get_ram():
##    global mem
##    global finish_tracking
##    name="PHSFService.exe"
##    mem=0.0
##    while(finish_tracking==1):
##        for proc in psutil.process_iter():
##            if( (proc.name()==name) & (mem<proc.memory_info()[-1]/(1024*1024))):
##                mem=proc.memory_info()[-1]/(1024*1024)
##        time.sleep(2)

def UpdateExcel(TestName,UpdateValues):
    global wb
    global testdict
    global report
    try:
        sheetname=testdict[TestName][0]
        print(sheetname)
        sheet1=wb[sheetname]
        print(sheet1)
        coltowrite=testdict[TestName][1]
        print("Column to write : {}".format(coltowrite))
    except Exception as Ex:
        print("Failed due to {}".format(Ex))

    rowtowrite=0
    for i in range(20,1000):
        rowtowrite=i
        if (sheet1.cell(row=i,column=1).value==None):
            break

    print("Values to be updated {}".format(len(UpdateValues)))    

    for data in UpdateValues:
        try:
            sheet1.cell(row=i,column=coltowrite).value=data
            coltowrite+=1
        except Exeption as e:
            print("\nAn Error occured while updating Excel Report. Error {}".format(e))
            logging.warning("\nAn Error occured while updating Excel Report. Error {}".format(e))
    wb.save(report)
            










def Perf_Counter_start(r):
    global finish_tracking
    try:
        k=sb.run([root_path+r"\bin\PhApi4.0_IntegrationTests\PerformanceCounterLogger.exe",root_path+r"\xmls\WritePerformanceCounters.xml",r,"1"])
    except Exception as e:
        print('Start Performance counter Error ',e)
        logging.warning("\nAn Error occured")
        finish_tracking=0
    while True:
        if (finish_tracking==0):  
            try:
                k=sb.run([root_path+r"\bin\PhApi4.0_IntegrationTests\PerformanceCounterLogger.exe","Dummy.xml",r,"0"])
                break
            except:
                print('Stop Performance counter Error')
                logging.warning("\nStop Performance counter Error")
                break

        time.sleep(5)
                
       

def Perf_Counter_stop():
    try:
        k=sb.run([root_path+r"\bin\PhApi4.0_IntegrationTests\PerformanceCounterLogger.exe","Dummy.xml","0"])
    except:
        print('Stop Performance counter Error')
        logging.warning("\nAn Error occured")
        
    

def Read_Max_RAM(FileLocation,TestName):
    df=pd.read_csv('D:\\WriteOutput\\'+TestName+r'.csv')
    df=df.iloc[1:,:]
    mn=df.describe().iloc[1,0]/(1024*1024)
    mx=df.describe().iloc[7,0]/(1024*1024)
    return([mn,mx])
    

def UpdateStartTime():
    
    s=''
    with open(r"C:\temp\PHbin\PhApi4.0_IntegrationTests\DataStartTime.txt","r+") as f:
        old=f.read()
        print(old)
        time=parse(old)
        newtime=(time + timedelta(hours=24))
        s=str(newtime)
        s=s.replace(" ","T",)
        
        s=s.replace("+",".0000000+",)
        print(s)
        f.seek(0)
        f.write(s)
        f.close()

    with open(r"C:\temp\PHbin\PhApi4.0_IntegrationTests\DataEndTime.txt","r+") as f:
        f.seek(0)
        f.write(s)
        f.close()


##def UpdateXml(loc,time1):
##    testfile=xml.dom.minidom.parse(loc)
##    print(testfile)
##   
##    ele=testfile.getElementsByTagName('QueryRangeStartOffsetInHours')[0]
##    
##    val=ele.firstChild.data
##    
##   
##    newval=int(val) + int((24-int(time1)))
##    print('New offset : {}'.format(newval))
##    ele.firstChild.data=str(newval)
##    f=open(loc,'w')
##    testfile.writexml(f)
##    f.close()
    

def runtestinloop(rp,r):
    
    global finish_tracking
    
    for i in range(int(r[5]),int(r[4])):
        try:
            t1=threading.Thread(target=Perf_Counter_start,args=(str(r[3]+'_'+str(i)),))
            t1.start()
            k=sb.run([rp+r[1],rp+r[2]])
            
        except Exception as ex:
            print('An Error {}'.format(ex))
            logging.warning("\nAn Error occured")
        logging.info("\nStarting WriteCalculations for day: {}".format(i))
        finish_tracking=0
        t1.join()
        time.sleep(5)
        WriteClaculations('D:\PHWriteNFRTestLogs',r[4])

        finish_tracking=1
        logging.info("\nRenaming Log files for test: {} ".format(r[4]))
        os.rename('D:\PHWriteNFRTestLogs','D:\Day'+str(i))
        UpdateStartTime()
                
            
        ##        logging.info("Sleeping for an hour")
        ##        print("Sleeping for an hour")
        ##        time.sleep(3600)
        logging.info("\nData Written for Day: {}".format(i))
##        while(True):
##            t2=time.time()
##            logging.info("\nHow Close to an hour: {}".format((t2-t1)%3600))
##            if(((t2-t1)%3600)<300):
##                logging.info("\nOffset Change: {}".format((t2-t1)/3600))
##                break
##            else:
##                logging.info("\n60 seconds of sleep")
##                time.sleep(60)
##        
##
##        UpdateXml(rp+r[2],(t2-t1)/3600) 
        
        

        
        
    
    
    
def ReplaceFiles(destination,source):
    shutil.copy(dst=destination,src=source)


def ReadCalculation(FileLocation,TestName):

    try:
        RAM=Read_Max_RAM(FileLocation,TestName)

    except Exception as Ex:
        print(Ex)
        logging.info("\nException during RAM Calculation: {} ".format(Ex))

    result=[]
    result.append(RAM[0])
    result.append(RAM[1])

    if('HLAS' in TestName):
        File='\AlarmReadResponse.csv'
        logging.info("\nAlarmReadResponse.csv file generated for test: {} ".format(TestName) )
    else:
        File='\TagReadResponse.csv'
        logging.info("\nTagReadResponse.csv file generated for test: {} ".format(TestName) )



    try:
        df=pd.read_csv(FileLocation+File,skiprows=1)
        df2=df.drop(['Version', 'ThreadId', 'Datetime', 'QueryRangeStart', 'QueryRangeEnd','BoundaryMode',
       'Iteration','ActualRowCount','ExpectedRowCount', 'Result', 'DesiredNoOfSamples'],axis=1)
        dfagg=df2.groupby(['TagCount','QueryRange(Minutes)']).agg('mean')
        arr=dfagg.iloc[0:,0:].values.reshape(-1,1)
        
        for i in arr:
            result.append(float(i))

    except Exception as E:
        print(E)
        logging.info("\n Exception {}".format(E))
        

    with open(FileLocation+'\ReadResults.csv',mode='w') as csvfile:
        wrtr=csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        wrtr.writerow(result)
        logging.info("\nConsolidated results generated in file for Test: {} ".format(TestName) )
    
    try:
        UpdateExcel(TestName,result)
    except Exception as e:
        logging.info("\nWriting results to excel failed due to: {} ".format(e))
        print("Writing to excel report failed due to {}".format(e))
        
    




    
def WriteClaculations(FileLocation,TestName):
    global report
    results=[]

    try:
        RAM=Read_Max_RAM(FileLocation,TestName)
        
    except Exception as Ex:
        print(Ex)
        logging.info("\nException: {} ".format(Ex))

    results.append(RAM[0])
    results.append(RAM[1])
    


    try:
        df=pd.read_csv(FileLocation+'\WriteLatency.csv',skiprows=1) #Reading CSV
        dftags=df[df['Data Kind']=='Tags'] #Segeregating Alarms and tags data
        dfalarms=df[df['Data Kind']=='Alarms']
        dftags=dftags.drop(['Version', 'ThreadId', 'Count', 'Datetime', 'WR Timestamp','Client Transaction Id', 'Write Request Id', 'Quantity', 'Data Kind'],axis=1) #discarding unnecessary data for this report
        dfalarms=dfalarms.drop(['Version', 'ThreadId', 'Count', 'Datetime', 'WR Timestamp','Client Transaction Id', 'Write Request Id', 'Quantity', 'Data Kind'],axis=1)
        alarmagg=dfalarms.describe()    #calculating aggregate
        tagagg=dftags.describe()
        LLindexes=[(3,0),(7,0),(1,0),(6,0),(3,1),(7,1),(1,1),(6,1)] #index for selection of required data as per deployment
        RLindexes=[(3,0),(7,0),(1,0),(6,0),(3,1),(7,1),(1,1),(6,1),(3,2),(7,2),(1,2),(6,2),(3,3),(7,3),(1,3),(6,3)]
        RLLTAindexes=[(3,0),(7,0),(1,0),(6,0),(3,1),(7,1),(1,1),(6,1),(3,2),(7,2),(1,2),(6,2),(3,3),(7,3),(1,3),(6,3),(3,4),(7,4),(1,4),(6,4),(3,5),(7,5),(1,5),(6,5)]
        LLLTAindexes=[(3,0),(7,0),(1,0),(6,0),(3,1),(7,1),(1,1),(6,1),(3,4),(7,4),(1,4),(6,4)]
    except Exception as E:
        print(E)
        logging.info("\n Exception {}".format(E))
    
    if 'LocalLog' in report:
        indexes=LLindexes
    elif 'Red' in report:
        indexes=RLindexes
    elif 'Red' in report and 'LTA' in report:
        indexes=RLLTAindexes
    elif 'LTA' in report:
        indexes=LLLTAindexes
    else:
        indexes=[]

    for idx in indexes:
        results.append(alarmagg.iloc[idx])
    for idx in indexes:
        results.append(tagagg.iloc[idx])

    UpdateExcel(TestName,results)

        
    with open(FileLocation+'\WriteResults.csv',mode='w') as csvfile:
        
        logging.info("\nGenerating WriteResults.csv for test: {} ".format(TestName) )
        
        try:
            wrtr=csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            wrtr.writerow(['For Report Std',results])

        except Exception as Ex:
            print(Ex)
            logging.info("\nException: {} ".format(Ex))

            
        
    print('Calculations finished')
    logging.info("\nCalculations finished for test: {} ".format(TestName) )




def SetRootPath(rp):
    global root_path
    root_path=rp
    logging.info("\nSetting root path as: {} ".format(rp) )
    PopulateDict(sys.argv[1])

def RenameStorage(rp,r):
    print("RestartingService")
    logging.info("\nRestarting PH StoreAndForeward Services")
    name= "PHSFService.exe"
    for proc in psutil.process_iter():
        if proc.name()==name:
            proc.kill()
            logging.info("\nKilling process")
    time.sleep(60)
    logging.info("\nDelay for 60 seconds")
    os.rename('C:\ProgramData\Siemens\PHCoreLog','C:\ProgramData\Siemens\\'+str(r[4]))
    logging.info("\nRenaming Log files for test: {} ".format(str(r[4])) )
    args=['sc','start','StoreAndForward']
    time.sleep(60)
    logging.info("\nDelay for 60 seconds")
    result=sb.run(args)

def DelayInMin(rp,r):
    time.sleep(int(r[1])*60)
    logging.info("\nDelay for 60 seconds")
    
def WriteAndSaveResults(rp,r):
    try:
        k=sb.run([rp+r[1],rp+r[2]])
    except:
        print('An Error')
        logging.warning("\nAn Error occured")
        finish_tracking=0
    logging.info("\nStarting WriteCalculations for test: {}".format(r[4]))
    WriteClaculations('D:\PHWriteNFRTestLogs',r[4])

    logging.info("\nRenaming Log files for test: {} ".format(r[4]) )
    os.rename('D:\PHWriteNFRTestLogs','D:\F'+r[4])
    
def ReadAndSaveResults(rp,r):
    k=sb.run([rp+r[1],rp+r[2]])
    logging.info("\nStarting ReadCalculations for test: {}".format(r[4]))
    ReadCalculation('D:\PHReadNFRTestLogs',r[4])

    logging.info("\nRenaming Log files for test: {} ".format(r[4]) )
    os.rename('D:\PHReadNFRTestLogs','D:\F'+r[4])

def RunExe(rp,r):
    print("RunnningEXE")
    logging.info("\nRunning .exe")
    print(rp+r[1])
    print(rp+r[2])
    k=sb.run([rp+r[1],rp+r[2]])
    


def MixedModeAndSave(rp,r):
    k=sb.run([rp+r[1],rp+r[2],rp+r[3]])
    logging.info("\nRenaming Log files for test: {} ".format(r[4]) )
    os.rename('D:\PHReadNFRTestLogs','D:\F'+r[4])

#0 to set root path,1 to write and save results,2 to read and save results,3 to Mixed mode and save results,4 to remove storage
#5 to add delay #6 to replace files #7 to run any exe with arguements #8 run test in loop
def SelectFn(r):
    global finish_tracking
    if r[0]=='0':
        SetRootPath(r[1])
    elif r[0]=='1':
        print(r[4])
        t1=threading.Thread(target=Perf_Counter_start,args=(r[4],))
        t1.start()
        try:
            WriteAndSaveResults(root_path,r)
        except Exception as e:
            print(e)
        
        finish_tracking=0
        t1.join()
        finish_tracking=1
    elif r[0]=='2':
        t1=threading.Thread(target=Perf_Counter_start,args=(r[4],))
        t1.start()
        try:
            ReadAndSaveResults(root_path,r)
        except Exception as e:
            print(e)
    
        finish_tracking=0
        t1.join()
        finish_tracking=1
    elif r[0]=='3':
        MixedModeAndSave(root_path,r)
    elif r[0]=='4':
        RenameStorage(root_path,r)
    elif r[0]=='5':
        print("starting Delay")
        DelayInMin(root_path,r)
        logging.info("\nStarting delay")
    elif r[0]=='6':
        print("Replacing files")
        ReplaceFiles(r[1],root_path+r[2])
        logging.info("\nReplacing Files")
    elif r[0]=='7':
        RunExe(root_path,r)
    elif r[0]=='8':
        runtestinloop(root_path,r)
    else:
        print("default")
        logging.info("\nDefault")

def PopulateDict(filename):
    global report
    global testdict
    global wb
    global root_path
    report=filename
    wb=openpyxl.load_workbook(filename)
    #load Test details used as map to update excel report
    with open(root_path + r'\New_Test\test_details.csv') as Details_CSV:
        csvrdr=csv.reader(Details_CSV,delimiter=',')
        for row in csvrdr:
            try:
                if(row[0]!='Testname'):
                    testdict[row[0]]=[row[1],int(row[2])]
            except Exception as e:
                print(e)
                logging.info("\nTest Details Dictionary Loaded")
                break

    

if __name__=="__main__":
        
    with open('test_seq.csv') as csv_file:
        csv_readr=csv.reader(csv_file,delimiter=',')
        for row in csv_readr:
            #print(type(row[0]),type(row[1]),type(row[2]),type(row[3]))
            logging.info("\nStarting for test: {} ".format(row) )
            SelectFn(row)


